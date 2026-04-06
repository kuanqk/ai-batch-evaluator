"""Parse CSV / Excel batch uploads (columns file_path, file_url)."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook

from pipeline.parser import parse_file_path


def _require_cols(headers: list[str]) -> None:
    hset = {h.strip().lower() for h in headers if h and str(h).strip()}
    if "file_path" not in hset or "file_url" not in hset:
        raise ValueError("Required columns: file_path, file_url")


def _parse_path_meta(file_path: str | None) -> dict[str, Any]:
    fp = (file_path or "").strip()
    if not fp:
        return {"city": None, "trainer": None, "group_name": None, "file_name": None}
    try:
        p = parse_file_path(fp)
        return {
            "city": p.city,
            "trainer": p.trainer,
            "group_name": p.group_name,
            "file_name": p.file_name,
        }
    except ValueError:
        return {"city": None, "trainer": None, "group_name": None, "file_name": None}


def parse_batch_upload(content: bytes, filename: str) -> tuple[list[dict[str, Any]], int]:
    """Returns (rows for bulk insert, skipped_row_count)."""
    name_lower = (filename or "").lower()
    skipped = 0
    rows: list[dict[str, Any]] = []

    if name_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header_row = next(it, None)
            if not header_row:
                raise ValueError("Empty Excel file")
            headers = [str(c).strip().lower() if c is not None else "" for c in header_row]
            _require_cols(headers)
            i_path = headers.index("file_path")
            i_url = headers.index("file_url")
            for row in it:
                if not row:
                    skipped += 1
                    continue
                fu = row[i_url] if i_url < len(row) else None
                fp = row[i_path] if i_path < len(row) else None
                fu_s = str(fu).strip() if fu is not None else ""
                if not fu_s:
                    skipped += 1
                    continue
                fp_s = str(fp).strip() if fp is not None else ""
                meta = _parse_path_meta(fp_s)
                meta["file_path"] = fp_s or None
                meta["file_url"] = fu_s
                rows.append(meta)
        finally:
            wb.close()
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV has no headers")
        norm_fields = [h.strip().lower() for h in reader.fieldnames if h]
        _require_cols(norm_fields)
        for raw in reader:
            row = {((k or "").strip().lower()): (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
            fu = (row.get("file_url") or "").strip()
            if not fu:
                skipped += 1
                continue
            fp = (row.get("file_path") or "").strip()
            meta = _parse_path_meta(fp)
            meta["file_path"] = fp or None
            meta["file_url"] = fu
            rows.append(meta)

    return rows, skipped
